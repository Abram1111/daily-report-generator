""import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime
import shutil
import os

st.title("📝 Daily Visit Report Generator")

# Load source file
df = pd.read_excel("Data base.xlsx")

visits = st.number_input("Number of visits", min_value=1, step=1)

filled_data = []

for i in range(visits):
    st.subheader(f"Visit {i + 1}")

    is_office = st.checkbox(f"Is Visit {i + 1} an Office Visit?", key=f"office_{i}")

    if is_office:
        filled_data.append([
            datetime.now().strftime('%Y-%m-%d'),
            "Office",  # Site Name
            "", "", "", "",  # Governorate, Address, Contact Name, Contact No.
            "", "", "",  # Type, Task Status, Task Type
            "", "", "",  # Model, SN, Work
            "", ""  # Travel, Tech Report, Case No.
        ])
    else:
        serial_input = st.text_input(f"Serial Number {i + 1}", key=f"serial_{i}")
        row = df[df["Serial Number"].astype(str) == serial_input]

        if not row.empty:
            row = row.iloc[0]
            model = str(row["Model"])

            st.markdown(f"**Customer:** {row['Customer Name']}")
            st.markdown(f"**Governorate:** {row['Governorate']}")
            st.markdown(f"**Address:** {row['Address']}")
            st.markdown(f"**Model:** {model}")

            has_printer = False
            printer_serial_input = ""
            printer_row = None
            if "CR" in model.upper():
                has_printer = st.checkbox(f"Does Visit {i + 1} have a printer with it?", key=f"printer_check_{i}")
                if has_printer:
                    printer_serial_input = st.text_input(f"Printer Serial Number for Visit {i + 1}", key=f"printer_sn_{i}")
                    printer_df = df[df["Serial Number"].astype(str) == printer_serial_input]
                    if not printer_df.empty:
                        printer_row = printer_df.iloc[0]
                    else:
                        st.warning("⚠️ Printer serial not found — only main device will be included.")

            task_type = st.text_input(f"Task Type {i+1}", key=f"task_{i}")
            task_status = st.selectbox(f"Task Status {i+1}", ["Complete", "NOT Complete"], key=f"status_{i}")
            visit_type = st.selectbox(f"Type {i+1}", ["PPM", "Service"], key=f"type_{i}")
            work = st.text_input(f"Total Working Time {i+1}", key=f"work_{i}")
            tech_report = st.text_input(f"Technical Report No. {i+1}", key=f"report_{i}")

            final_sn = serial_input
            final_model = model

            if has_printer and printer_row is not None:
                final_sn = f"{serial_input}, {printer_serial_input}"
                final_model = f"{model}, {printer_row['Model']}"

            filled_data.append([
                datetime.now().strftime('%Y-%m-%d'),
                row["Customer Name"],
                row["Governorate"],
                row["Address"],
                row["Contact Person 1"],
                row["Contact Number 1"],
                visit_type,
                task_status,
                task_type,
                final_model,
                final_sn,
                work,
                "",
                tech_report,
                ""
            ])
        else:
            st.warning("❌ Serial not found.")
            add_manual = st.checkbox(f"Add this new device manually?", key=f"manual_add_{i}")
            if add_manual:
                customer = st.text_input(f"Customer Name {i + 1}", key=f"cust_{i}")
                gov = st.text_input(f"Governorate {i + 1}", key=f"gov_{i}")
                address = st.text_input(f"Address {i + 1}", key=f"addr_{i}")
                contact_name = st.text_input(f"Contact Person {i + 1}", key=f"cp_{i}")
                contact_no = st.text_input(f"Contact No. {i + 1}", key=f"cno_{i}")
                model = st.text_input(f"Model {i + 1}", key=f"model_{i}")

                task_type = st.text_input(f"Task Type {i+1}", key=f"task_m_{i}")
                task_status = st.selectbox(f"Task Status {i+1}", ["Complete", "NOT Complete"], key=f"status_m_{i}")
                visit_type = st.selectbox(f"Type {i+1}", ["PPM", "Service"], key=f"type_m_{i}")
                work = st.text_input(f"Total Working Time {i+1}", key=f"work_m_{i}")
                tech_report = st.text_input(f"Technical Report No. {i+1}", key=f"report_m_{i}")

                filled_data.append([
                    datetime.now().strftime('%Y-%m-%d'),
                    customer,
                    gov,
                    address,
                    contact_name,
                    contact_no,
                    visit_type,
                    task_status,
                    task_type,
                    model,
                    serial_input,
                    work,
                    "",
                    tech_report,
                    ""
                ])

                # Save new entry to Excel
                new_row = {
                    "Customer Name": customer,
                    "Governorate": gov,
                    "Address": address,
                    "Contact Person 1": contact_name,
                    "Contact Number 1": contact_no,
                    "Model": model,
                    "Serial Number": serial_input
                }
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                df.to_excel("Data base.xlsx", index=False)
                st.success("✅ New device added to database!")

# When ready to export
if st.button("Generate Excel Report"):
    if filled_data:
        template = "Daily Report Form.xlsx"
        output = f"filled_visits_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        shutil.copy(template, output)
        wb = load_workbook(output)
        ws = wb.active

        for i, row_data in enumerate(filled_data, start=2):
            for j, val in enumerate(row_data, start=2):
                cell = ws.cell(row=i, column=j, value=val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)

        wb.save(output)
        st.success("✅ File generated!")
        with open(output, "rb") as f:
            st.download_button("📥 Download Report", f, file_name=output)
    else:
        st.warning("No data to export.")
